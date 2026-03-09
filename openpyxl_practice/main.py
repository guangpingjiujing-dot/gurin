"""
openpyxlの基本動作を学習するサンプルプログラム

各トピックごとに関数を作成しています。
main関数で実行したいトピックのコメントアウトを外して実行してください。
"""

import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Excelファイルを保存するディレクトリ
EXCEL_DIR = Path(__file__).parent / "excel_files"
EXCEL_DIR.mkdir(exist_ok=True)

def topic1_create_workbook():
    """トピック1: 基本的なワークブックの作成と保存"""
    print("\n=== トピック1: 基本的なワークブックの作成と保存 ===")
    
    # 新しいワークブックを作成
    wb = Workbook()
    
    # アクティブなワークシートを取得
    ws = wb.active
    ws.title = "サンプルシート"
    
    # セルに値を設定
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    ws["A2"] = 100
    ws["B2"] = 200
    
    # ファイルを保存
    file_path = EXCEL_DIR / "topic1_basic_workbook.xlsx"
    wb.save(file_path)
    print(f"ワークブックを作成して保存しました: {file_path}")
    
    wb.close()


def topic2_read_cell_values():
    """トピック2: セルからの値の読み取り"""
    print("\n=== トピック2: セルからの値の読み取り ===")
    
    # まずサンプルファイルを作成
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "名前"
    ws["B1"] = "年齢"
    ws["A2"] = "太郎"
    ws["B2"] = 25
    ws["A3"] = "花子"
    ws["B3"] = 30
    
    file_path = EXCEL_DIR / "topic2_read_sample.xlsx"
    wb.save(file_path)
    wb.close()
    
    # ファイルを読み込む
    wb = load_workbook(file_path)
    ws = wb.active
    
    # セルの値を読み取る方法
    print(f"A1の値: {ws['A1'].value}")
    print(f"B2の値: {ws['B2'].value}")
    print(f"A3の値: {ws.cell(row=3, column=1).value}")
    
    # 範囲を読み取る
    print("\n範囲A1:B3の値:")
    for row in ws["A1:B3"]:
        for cell in row:
            print(f"  {cell.coordinate}: {cell.value}")
    
    wb.close()


def topic3_write_cells():
    """トピック3: セルへの値の書き込み（様々な方法）"""
    print("\n=== トピック3: セルへの値の書き込み ===")
    
    wb = Workbook()
    ws = wb.active
    
    # 方法1: セル参照で直接書き込み
    ws["A1"] = "方法1: セル参照"
    ws["A2"] = 123
    
    # 方法2: cell()メソッドを使用
    ws.cell(row=1, column=2, value="方法2: cell()メソッド")
    ws.cell(row=2, column=2, value=456)
    
    # 方法3: 行ごとに書き込み
    ws.append(["方法3: append()", "データ1", "データ2"])
    ws.append([None, "データ3", "データ4"])
    
    # 方法4: 範囲に一括書き込み
    data = [
        ["商品名", "価格", "数量"],
        ["りんご", 100, 5],
        ["バナナ", 80, 3],
        ["オレンジ", 120, 2]
    ]
    for row_idx, row_data in enumerate(data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = EXCEL_DIR / "topic3_write_cells.xlsx"
    wb.save(file_path)
    print(f"セルへの書き込みサンプルを保存しました: {file_path}")
    
    wb.close()


def topic4_row_column_operations():
    """トピック4: 行・列の操作"""
    print("\n=== トピック4: 行・列の操作 ===")
    
    wb = Workbook()
    ws = wb.active
    cell
    # データを準備
    for row in range(1, 6):
        for col in range(1, 4):
            ws.cell(row=row, column=col, value=f"R{row}C{col}")
    
    # 行の高さを設定
    ws.row_dimensions[1].height = 30
    
    # 列の幅を設定
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    
    # 行を挿入（2行目の前に1行挿入）
    ws.insert_rows(2)
    ws["A2"] = "挿入された行"
    
    # 列を挿入（B列の前に1列挿入）
    ws.insert_cols(2)
    ws["B1"] = "挿入された列"
    
    # 行を削除（例: 5行目を削除）
    # ws.delete_rows(5)
    
    # 列を削除（例: D列を削除）
    # ws.delete_cols(4)
    
    file_path = EXCEL_DIR / "topic4_row_column.xlsx"
    wb.save(file_path)
    print(f"行・列操作のサンプルを保存しました: {file_path}")
    
    wb.close()


def topic5_cell_styles():
    """トピック5: セルのスタイル設定"""
    print("\n=== トピック5: セルのスタイル設定 ===")
    
    wb = Workbook()
    ws = wb.active
    
    # ヘッダー行のスタイル
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws["A1"] = "商品名"
    ws["B1"] = "価格"
    ws["C1"] = "在庫"
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # データ行
    data = [
        ["りんご", 100, 50],
        ["バナナ", 80, 30],
        ["オレンジ", 120, 20]
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 価格列（B列）を右揃え
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="right")
            
            # 在庫が20以下の場合は背景色を変更
            if col_idx == 3 and value <= 20:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 罫線を設定
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws["A1:C4"]:
        for cell in row:
            cell.border = thin_border
    
    # 行の高さと列の幅を調整
    ws.row_dimensions[1].height = 25
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    
    file_path = EXCEL_DIR / "topic5_styles.xlsx"
    wb.save(file_path)
    print(f"スタイル設定のサンプルを保存しました: {file_path}")
    
    wb.close()


def topic6_worksheet_operations():
    """トピック6: ワークシートの操作"""
    print("\n=== トピック6: ワークシートの操作 ===")
    
    wb = Workbook()
    
    # デフォルトのシートを削除
    wb.remove(wb.active)
    
    # 新しいシートを作成
    ws1 = wb.create_sheet("売上データ")
    ws2 = wb.create_sheet("在庫データ")
    ws3 = wb.create_sheet("顧客データ", 0)  # 0番目に挿入（最初に配置）
    
    # 各シートにデータを追加
    ws1["A1"] = "売上シート"
    ws1["A2"] = "2024年1月: 100万円"
    
    ws2["A1"] = "在庫シート"
    ws2["A2"] = "りんご: 50個"
    
    ws3["A1"] = "顧客シート"
    ws3["A2"] = "顧客数: 100名"
    
    # シート名を変更
    ws1.title = "売上管理"
    
    # シートの順序を確認
    print("シートの順序:")
    for idx, sheet in enumerate(wb.worksheets, start=1):
        print(f"  {idx}. {sheet.title}")
    
    # シートを削除（コメントアウト中）
    # wb.remove(ws2)
    
    file_path = EXCEL_DIR / "topic6_worksheets.xlsx"
    wb.save(file_path)
    print(f"ワークシート操作のサンプルを保存しました: {file_path}")
    
    wb.close()


def topic7_formulas():
    """トピック7: 数式の使用"""
    print("\n=== トピック7: 数式の使用 ===")
    
    wb = Workbook()
    ws = wb.active
    
    # データを設定
    ws["A1"] = "商品名"
    ws["B1"] = "単価"
    ws["C1"] = "数量"
    ws["D1"] = "合計"
    
    data = [
        ["りんご", 100, 5],
        ["バナナ", 80, 3],
        ["オレンジ", 120, 2]
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 数式を設定（合計 = 単価 × 数量）
    for row in range(2, 5):
        ws[f"D{row}"] = f"=B{row}*C{row}"
    
    # 合計行を追加
    ws["A5"] = "合計"
    ws["D5"] = "=SUM(D2:D4)"
    
    # 平均を計算
    ws["A6"] = "平均"
    ws["D6"] = "=AVERAGE(D2:D4)"
    
    file_path = EXCEL_DIR / "topic7_formulas.xlsx"
    wb.save(file_path)
    print(f"数式のサンプルを保存しました: {file_path}")
    
    # 計算結果を確認
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    print(f"\n計算結果:")
    print(f"  D2の値: {ws['D2'].value}")
    print(f"  D5の値（合計）: {ws['D5'].value}")
    print(f"  D6の値（平均）: {ws['D6'].value}")
    
    wb.close()


def topic8_merge_cells():
    """トピック8: セルの結合"""
    print("\n=== トピック8: セルの結合 ===")
    
    wb = Workbook()
    ws = wb.active
    
    # セルを結合
    ws.merge_cells("A1:D1")
    ws["A1"] = "売上管理表"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=14)
    
    # ヘッダー行
    headers = ["商品名", "単価", "数量", "合計"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # データ行
    data = [
        ["りんご", 100, 5, 500],
        ["バナナ", 80, 3, 240],
        ["オレンジ", 120, 2, 240]
    ]
    
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 合計行でセルを結合
    ws.merge_cells("A6:B6")
    ws["A6"] = "合計"
    ws["D6"] = "=SUM(D3:D5)"
    
    # 列の幅を調整
    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 12
    
    file_path = EXCEL_DIR / "topic8_merge_cells.xlsx"
    wb.save(file_path)
    print(f"セル結合のサンプルを保存しました: {file_path}")
    
    wb.close()


def topic9_iterate_data():
    """トピック9: データの反復処理"""
    print("\n=== トピック9: データの反復処理 ===")
    
    wb = Workbook()
    ws = wb.active
    
    # サンプルデータを作成
    data = [
        ["ID", "名前", "年齢", "部署"],
        [1, "太郎", 25, "営業"],
        [2, "花子", 30, "開発"],
        [3, "次郎", 28, "営業"],
        [4, "美咲", 32, "開発"],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # 方法1: 行ごとに反復
    print("\n方法1: 行ごとに反復")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        print(f"  {row}")
    
    # 方法2: 列ごとに反復
    print("\n方法2: 列ごとに反復（1列目のみ）")
    for cell in ws.iter_cols(min_col=1, max_col=1, min_row=2, values_only=True):
        print(f"  {list(cell)}")
    
    # 方法3: 特定の範囲を反復
    print("\n方法3: 特定の範囲を反復（名前と年齢のみ）")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3, values_only=True):
        print(f"  名前: {row[0]}, 年齢: {row[1]}")
    
    # 方法4: 条件に合うデータを抽出
    print("\n方法4: 条件に合うデータを抽出（年齢が30以上）")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[2] and row[2] >= 30:  # 年齢が30以上
            print(f"  {row[1]}さん（{row[2]}歳）")
    
    file_path = EXCEL_DIR / "topic9_iterate.xlsx"
    wb.save(file_path)
    print(f"\nデータ反復処理のサンプルを保存しました: {file_path}")
    
    wb.close()


def topic10_comprehensive_example():
    """トピック10: 総合的な例（実践的なサンプル）"""
    print("\n=== トピック10: 総合的な例 ===")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "売上レポート"
    
    # タイトル行
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "2024年1月 売上レポート"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # ヘッダー行
    headers = ["日付", "商品名", "単価", "数量", "小計"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[2].height = 25
    
    # データ行
    sales_data = [
        ["2024/1/1", "りんご", 100, 10],
        ["2024/1/2", "バナナ", 80, 15],
        ["2024/1/3", "オレンジ", 120, 8],
        ["2024/1/4", "りんご", 100, 12],
        ["2024/1/5", "バナナ", 80, 20],
    ]
    
    for row_idx, row_data in enumerate(sales_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 日付列を左揃え、数値列を右揃え
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")
        
        # 小計列に数式を設定
        ws[f"E{row_idx}"] = f"=C{row_idx}*D{row_idx}"
    
    # 合計行
    total_row = ws.max_row + 1
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws[f"A{total_row}"] = "合計"
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right")
    ws[f"E{total_row}"] = f"=SUM(E3:E{total_row-1})"
    ws[f"E{total_row}"].font = Font(bold=True)
    
    # 罫線を設定
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws[f"A2:E{total_row}"]:
        for cell in row:
            cell.border = thin_border
    
    # 列の幅を調整
    column_widths = {"A": 15, "B": 15, "C": 12, "D": 12, "E": 12}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    file_path = EXCEL_DIR / "topic10_comprehensive.xlsx"
    wb.save(file_path)
    print(f"総合的なサンプルを保存しました: {file_path}")
    
    wb.close()


def main():
    """メイン関数: 実行したいトピックのコメントアウトを外してください"""
    
    print("openpyxlの基本動作を学習するサンプルプログラム")
    print("=" * 60)
    
    # 実行したいトピックのコメントアウトを外してください
    
    # topic1_create_workbook()           # 基本的なワークブックの作成と保存
    topic2_read_cell_values()          # セルからの値の読み取り
    # topic3_write_cells()               # セルへの値の書き込み
    # topic4_row_column_operations()     # 行・列の操作
    # topic5_cell_styles()               # セルのスタイル設定
    # topic6_worksheet_operations()      # ワークシートの操作
    # topic7_formulas()                  # 数式の使用
    # topic8_merge_cells()               # セルの結合
    # topic9_iterate_data()              # データの反復処理
    # topic10_comprehensive_example()    # 総合的な例
    
    print("\n" + "=" * 60)
    print("実行完了！excel_filesフォルダに生成されたExcelファイルを確認してください。")


if __name__ == "__main__":
    main()
